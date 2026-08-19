from io import BytesIO

from openpyxl import load_workbook

from agente_qa.export.excel import create_excel, AZURE_COLUMNS, MATRIZ_COLUMNS


def _sample():
    return {
        "TEST_CASES": [
            {
                "ID": "CP-AC-COT-00001",
                "Title": "CP-AC-COT-00001 Validar cotización|principal",
                "Product": "Cotizadores Web",
                "Module": "Cotizador autos colectivos",
                "Description": "Validar flujo principal.\n\nSegundo párrafo.",
                "Expected Result": "La cotización se genera correctamente.",
                "Preconditions": "Usuario autenticado.",
                "Related Use Case": "CU-001 Cotizar",
                "Steps": [
                    {"Step #": 1, "Action": "Ingresar datos", "Expected value": "Los datos son aceptados"},
                    {"Step #": 2, "Action": "Continuar", "Expected value": "Se muestra el resultado"},
                ],
            }
        ]
    }


def test_excel_has_expected_columns_and_sheets():
    blob = create_excel(_sample())
    workbook = load_workbook(BytesIO(blob), read_only=True)

    assert workbook.sheetnames == ["Azure Import", "Matriz QA"]
    assert list(workbook["Azure Import"].values)[0] == tuple(AZURE_COLUMNS)
    assert list(workbook["Matriz QA"].values)[0] == tuple(MATRIZ_COLUMNS)


def test_excel_preserves_one_header_row_and_separate_steps():
    blob = create_excel(_sample())
    workbook = load_workbook(BytesIO(blob), read_only=True)
    rows = list(workbook["Azure Import"].values)

    assert rows[1][1] == "Test Case"
    assert rows[1][2].startswith("CP-AC-COT-00001 ")
    assert rows[2][4] == 1
    assert rows[3][4] == 2
    assert rows[2][2] in (None, "")


def test_excel_removes_pipes_and_keeps_project_defaults():
    blob = create_excel(_sample())
    workbook = load_workbook(BytesIO(blob), read_only=True)
    rows = list(workbook["Azure Import"].values)
    header = rows[0]
    header_index = {name: i for i, name in enumerate(header)}

    for row in rows[1:]:
        assert "|" not in " ".join(str(value or "") for value in row)

    assert rows[1][header_index["Area Path"]] == r"COTIZADORES WEB\DESARROLLO"
    assert rows[1][header_index["Tipo Origen Proyecto"]] == "Proyecto"
